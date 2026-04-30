"""历史报价单 Excel/PDF 导入器。"""
from __future__ import annotations

import json
import os
import re
import uuid
import zipfile
from base64 import b64decode
from collections import defaultdict
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any, Callable
from xml.etree import ElementTree as ET

import httpx
import yaml
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from modules.m08_stats.schema import (
    HistoricalQuote,
    HistoricalQuoteCreate,
    QuoteFolderImportRequest,
    QuoteImportImage,
    QuoteImportRequest,
    QuoteImportResult,
    UploadedQuoteFileImportRequest,
)
from utils.secrets import load_from_keychain

ROOT = Path(__file__).resolve().parent.parent.parent
CONFIG_PATH = ROOT / "config.yaml"
DRAWING_NS = {
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
}

UpsertQuote = Callable[[HistoricalQuoteCreate], HistoricalQuote]


def load_quote_import_config() -> dict[str, Any]:
    """读取导入配置。"""
    if not CONFIG_PATH.exists():
        return {}
    return (yaml.safe_load(CONFIG_PATH.read_text(encoding="utf-8")) or {}).get("quote_import", {})


def import_quote_file(request: QuoteImportRequest, upsert_quote: UpsertQuote) -> QuoteImportResult:
    """导入 Excel/PDF 历史报价单，归一化为 M-08 历史报价。"""
    cfg = load_quote_import_config()
    usd_to_cny = float(request.usd_to_cny or cfg.get("usd_to_cny") or 7.0)
    source = _resolve_path(request.file_path)
    if not source.exists():
        raise FileNotFoundError(f"报价单文件不存在: {source}")

    import_id = _import_id(source)
    image_dir = _resolve_path(request.image_output_dir or cfg.get("image_output_dir") or "./data/imported_quote_images")
    suffix = source.suffix.lower()
    warnings: list[str] = []
    images: list[QuoteImportImage] = []
    if suffix in {".xlsx", ".xlsm"}:
        items, meta, images = _parse_xlsx(source, import_id, image_dir, usd_to_cny, request, warnings)
        file_type = "xlsx"
    elif suffix == ".pdf":
        items, meta = _parse_pdf(source, import_id, usd_to_cny, request, warnings)
        file_type = "pdf"
    else:
        raise ValueError(f"暂不支持的报价单格式: {source.suffix}")

    _maybe_enrich_with_llm(items, meta, request, cfg, warnings)

    quotes: list[HistoricalQuote] = []
    preview_items: list[dict[str, Any]] = []
    skipped = 0
    for item in items[:request.limit]:
        try:
            if not _item_has_price(item):
                preview_items.append(_item_to_preview_quote(item, meta, request))
                continue
            quote = _item_to_quote(item, meta, usd_to_cny, request)
            if request.import_to_db:
                quotes.append(upsert_quote(quote))
            else:
                quotes.append(_preview_quote(quote))
        except Exception as exc:  # noqa: BLE001 单行失败不影响整份文件
            skipped += 1
            warnings.append(f"跳过第 {item.get('row_index', '?')} 行: {exc}")

    return QuoteImportResult(
        import_id=import_id,
        source_file=str(source),
        file_type=file_type,
        quote_no=str(meta.get("quote_no") or ""),
        quote_date=str(meta.get("quote_date") or ""),
        customer_name=str(meta.get("customer_name") or ""),
        usd_to_cny=usd_to_cny,
        imported_count=len(quotes),
        recognized_count=len(quotes) + len(preview_items),
        preview_count=len(preview_items),
        skipped_count=skipped,
        warnings=warnings,
        images=images,
        quotes=quotes,
        preview_items=preview_items,
    )


def import_uploaded_quote_file(
    request: UploadedQuoteFileImportRequest,
    upsert_quote: UpsertQuote,
) -> QuoteImportResult:
    """保存浏览器上传文件，然后复用本地文件导入流程。"""
    cfg = load_quote_import_config()
    upload_dir = _resolve_path(cfg.get("upload_dir") or "./data/uploaded_quote_files")
    safe_relative = _safe_relative_path(request.relative_path or request.file_name)
    target = upload_dir / safe_relative
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_bytes(b64decode(request.content_base64))

    import_request = QuoteImportRequest(
        **request.model_dump(
            exclude={"content_base64", "file_name", "relative_path", "file_path"},
            exclude_none=True,
        ),
        file_path=str(target),
    )
    return import_quote_file(import_request, upsert_quote)


def import_quote_folder(
    request: QuoteFolderImportRequest,
    upsert_quote: UpsertQuote,
) -> dict[str, Any]:
    """递归导入本机文件夹内的报价单，给人工校验页返回统一汇总。"""
    folder = _resolve_path(request.folder_path).expanduser().resolve()
    if not folder.exists():
        raise FileNotFoundError(f"报价单文件夹不存在: {folder}")
    if not folder.is_dir():
        raise NotADirectoryError(f"不是文件夹路径: {folder}")

    quote_files = _quote_files(folder, request.recursive)[:request.max_files]
    warnings: list[str] = []
    if not quote_files:
        warnings.append("文件夹内未找到 .xlsx/.xlsm/.pdf 报价单")

    aggregate: dict[str, Any] = {
        "import_id": _import_id(folder),
        "source_file": str(folder),
        "file_type": "folder",
        "quote_no": "批量导入",
        "quote_date": "",
        "customer_name": "多客户",
        "usd_to_cny": float(request.usd_to_cny or load_quote_import_config().get("usd_to_cny") or 7.0),
        "imported_count": 0,
        "recognized_count": 0,
        "preview_count": 0,
        "skipped_count": 0,
        "warnings": warnings,
        "images": [],
        "quotes": [],
        "preview_items": [],
        "files": [],
        "found_count": len(quote_files),
    }

    for file_path in quote_files:
        relative = str(file_path.relative_to(folder))
        try:
            import_request = QuoteImportRequest(
                file_path=str(file_path),
                import_to_db=request.import_to_db,
                use_llm=request.use_llm,
                llm_provider_id=request.llm_provider_id,
                llm_model=request.llm_model,
                usd_to_cny=request.usd_to_cny,
                image_output_dir=request.image_output_dir,
                default_customer_tier=request.default_customer_tier,
                default_region=request.default_region,
                default_currency=request.default_currency,
                source_label=request.source_label,
                limit=request.limit,
                note=request.note,
            )
            result = import_quote_file(import_request, upsert_quote)
            aggregate["imported_count"] += result.imported_count
            aggregate["recognized_count"] += result.recognized_count
            aggregate["preview_count"] += result.preview_count
            aggregate["skipped_count"] += result.skipped_count
            aggregate["warnings"].extend(result.warnings)
            aggregate["images"].extend(image.model_dump(mode="json") for image in result.images)
            for quote in result.quotes:
                row = quote.model_dump(mode="json")
                row["_sourceFile"] = relative
                row["_quoteNo"] = result.quote_no
                aggregate["quotes"].append(row)
            for preview in result.preview_items:
                row = dict(preview)
                row["_sourceFile"] = relative
                row["_quoteNo"] = result.quote_no
                aggregate["preview_items"].append(row)
            aggregate["files"].append({
                "file_path": str(file_path),
                "relative_path": relative,
                "quote_no": result.quote_no,
                "customer_name": result.customer_name,
                "imported_count": result.recognized_count,
                "priced_count": result.imported_count,
                "preview_count": result.preview_count,
                "skipped_count": result.skipped_count,
                "image_count": len(result.images),
            })
        except Exception as exc:  # noqa: BLE001 单文件失败不影响整批导入
            aggregate["skipped_count"] += 1
            aggregate["warnings"].append(f"{relative}: {exc}")
            aggregate["files"].append({
                "file_path": str(file_path),
                "relative_path": relative,
                "status": "error",
                "error": str(exc),
                "imported_count": 0,
                "skipped_count": 1,
                "image_count": 0,
            })
    return aggregate


def analyze_price_factors(quotes: list[HistoricalQuote], min_support: int = 1, limit: int = 50) -> dict[str, Any]:
    """基于历史报价做轻量因素价格影响分析。"""
    values: dict[tuple[str, str], list[HistoricalQuote]] = defaultdict(list)
    for quote in quotes:
        factors = _analysis_factors(quote)
        for key, value in factors.items():
            if value not in (None, "", [], {}):
                values[(key, str(value))].append(quote)

    overall = _avg([quote.final_price_cny for quote in quotes])
    rows: list[dict[str, Any]] = []
    for (factor_key, factor_value), grouped in values.items():
        if len(grouped) < min_support:
            continue
        avg_price = _avg([quote.final_price_cny for quote in grouped])
        rows.append({
            "factor_key": factor_key,
            "factor_value": factor_value,
            "support_count": len(grouped),
            "avg_unit_price_cny": round(avg_price, 2),
            "min_unit_price_cny": round(min(quote.final_price_cny for quote in grouped), 2),
            "max_unit_price_cny": round(max(quote.final_price_cny for quote in grouped), 2),
            "impact_vs_overall": round(avg_price - overall, 2) if overall else 0,
            "sample_quote_ids": [quote.quote_id for quote in grouped[:5]],
        })
    rows.sort(key=lambda item: (abs(item["impact_vs_overall"]), item["support_count"]), reverse=True)
    return {
        "quote_count": len(quotes),
        "overall_avg_unit_price_cny": round(overall, 2) if overall else 0,
        "factors": rows[:limit],
    }


def _parse_xlsx(
    path: Path,
    import_id: str,
    image_dir: Path,
    usd_to_cny: float,
    request: QuoteImportRequest,
    warnings: list[str],
) -> tuple[list[dict[str, Any]], dict[str, Any], list[QuoteImportImage]]:
    workbook = load_workbook(path, data_only=False)
    meta = _workbook_meta(workbook, path)
    if request.default_currency:
        meta["currency"] = request.default_currency
    images = _extract_xlsx_images(path, import_id, image_dir, warnings)
    images_by_sheet_row: dict[tuple[str, int], list[QuoteImportImage]] = defaultdict(list)
    for image in images:
        images_by_sheet_row[(image.source_sheet, image.anchor_row)].append(image)

    items: list[dict[str, Any]] = []
    for sheet_index, ws in enumerate(workbook.worksheets, start=1):
        header_row = _detect_header_row(ws)
        if not header_row:
            continue
        columns = _build_column_map(ws, header_row)
        for row_index in range(header_row + 1, ws.max_row + 1):
            item = _row_item(ws, row_index, columns, meta, usd_to_cny)
            if not item:
                continue
            item["sheet_name"] = ws.title
            item["sheet_index"] = sheet_index
            item["row_index"] = row_index
            row_images = _row_image_paths(images_by_sheet_row.get((ws.title, row_index), []), columns)
            if row_images:
                item["image_paths"] = row_images
            items.append(item)
    if not items:
        warnings.append("未识别到有效产品行，请检查报价单表头或启用大模型辅助解析")
    return items, meta, images


def _parse_pdf(
    path: Path,
    import_id: str,
    usd_to_cny: float,
    request: QuoteImportRequest,
    warnings: list[str],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    try:
        from pypdf import PdfReader  # type: ignore
    except ImportError as exc:  # pragma: no cover - 仅 PDF 导入时触发
        raise RuntimeError("PDF 导入需要安装 pypdf") from exc

    reader = PdfReader(str(path))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    meta = _text_meta(text, path)
    if request.default_currency:
        meta["currency"] = request.default_currency
    if not meta.get("currency"):
        meta["currency"] = "USD" if re.search(r"\bUSD\b|\$", text, re.I) else "CNY"

    items: list[dict[str, Any]] = []
    for index, line in enumerate(text.splitlines(), start=1):
        item = _line_item(line, index, meta, usd_to_cny)
        if item:
            item["sheet_name"] = "PDF"
            item["sheet_index"] = 1
            item["row_index"] = index
            items.append(item)
    if not items:
        warnings.append("PDF 文本未直接识别到产品价格行，可启用大模型辅助解析")
    meta["raw_text_preview"] = text[:4000]
    meta["import_id"] = import_id
    return items, meta


def _workbook_meta(workbook, path: Path) -> dict[str, Any]:
    text_cells: list[tuple[str, int, int, Any]] = []
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value not in (None, ""):
                    text_cells.append((ws.title, cell.row, cell.column, cell.value))
    meta = _cells_meta(text_cells, path)
    joined = "\n".join(str(item[3]) for item in text_cells)
    meta["currency"] = _detect_currency(joined, None)
    return meta


def _cells_meta(cells: list[tuple[str, int, int, Any]], path: Path) -> dict[str, Any]:
    quote_no = ""
    quote_date = ""
    customer_name = ""
    for _sheet, _row, _col, value in cells:
        text = _clean_text(value)
        if "销售合同号" in text:
            quote_no = _extract_quote_no(text) or quote_no
            continue
        if not quote_no:
            quote_no = _extract_quote_no(text)
        if not quote_date:
            quote_date = _extract_date(value, text)
        if not customer_name:
            customer_name = _extract_customer(text)

    by_pos = {(sheet, row, col): value for sheet, row, col, value in cells}
    for sheet, row, col, value in cells:
        text = _clean_text(value)
        if re.search(r"^(no|编号|报价编号)\b|No\s*:|销售合同号", text, re.I):
            right = by_pos.get((sheet, row, col + 1))
            if right not in (None, ""):
                candidate = _clean_text(right)
                if _looks_like_quote_no(candidate):
                    quote_no = candidate
            elif "销售合同号" in text:
                quote_no = _extract_quote_no(text) or quote_no
        if re.search(r"^(to|客户名称|客户)\b|To\s*:", text, re.I):
            right = by_pos.get((sheet, row, col + 1))
            if right not in (None, ""):
                customer_name = _clean_text(right) or customer_name
        if re.search(r"date|日期", text, re.I):
            right = by_pos.get((sheet, row, col + 1))
            if right not in (None, ""):
                quote_date = _extract_date(right, _clean_text(right)) or quote_date

    return {
        "source_file": str(path),
        "quote_no": quote_no or path.stem,
        "quote_date": quote_date or _date_from_filename(path.name),
        "customer_name": customer_name,
    }


def _text_meta(text: str, path: Path) -> dict[str, Any]:
    return {
        "source_file": str(path),
        "quote_no": _extract_quote_no(text) or path.stem,
        "quote_date": _extract_date("", text) or _date_from_filename(path.name),
        "customer_name": _extract_customer(text),
    }


def _detect_header_row(ws) -> int | None:
    best_row = None
    best_score = 0
    for row in range(1, min(ws.max_row, 40) + 1):
        score = 0
        for col in range(1, ws.max_column + 1):
            header = _header_text(ws, row, col)
            if _classify_header(header):
                score += 1
            if _is_dimension_group(header):
                score += 2
        if score > best_score:
            best_score = score
            best_row = row
    return best_row if best_score >= 4 else None


def _build_column_map(ws, header_row: int) -> dict[str, Any]:
    columns: dict[str, Any] = {}
    for col in range(1, ws.max_column + 1):
        header = _header_text(ws, header_row, col)
        kind = _classify_header(header)
        if kind and kind not in columns:
            columns[kind] = col
        if kind in {"unit_price", "amount"}:
            columns.setdefault(f"{kind}_candidates", []).append(col)
        if kind == "image_col":
            columns.setdefault("image_col", col)
        if _is_split_dimension_group(header):
            columns.setdefault("width_cm", col)
            columns.setdefault("depth_cm", col + 1)
            columns.setdefault("height_cm", col + 2)
    return columns


def _row_item(ws, row_index: int, columns: dict[str, Any], meta: dict[str, Any], usd_to_cny: float) -> dict[str, Any] | None:
    raw_values = {
        get_column_letter(col): ws.cell(row_index, col).value
        for col in range(1, ws.max_column + 1)
        if ws.cell(row_index, col).value not in (None, "")
    }
    if not raw_values:
        return None

    product_name = _clean_text(_cell(ws, row_index, columns.get("product_name"))) or _clean_text(_cell(ws, row_index, columns.get("category_text")))
    item_no = _clean_text(_cell(ws, row_index, columns.get("item_no")))
    material = _clean_text(_cell(ws, row_index, columns.get("material")))
    finish = _clean_text(_cell(ws, row_index, columns.get("finish"))) or _clean_text(_cell(ws, row_index, columns.get("color")))
    fabric = _clean_text(_cell(ws, row_index, columns.get("fabric")))
    requirements = _clean_text(_cell(ws, row_index, columns.get("requirements")))
    unit_text = _clean_text(_cell(ws, row_index, columns.get("unit")))
    category_text = product_name or _clean_text(_cell(ws, row_index, columns.get("category_text"))) or material
    category = _category_code(category_text, item_no)
    currency = _detect_currency(" ".join(str(value) for value in raw_values.values()), meta.get("currency")) or "CNY"
    unit_price = _first_float_in_columns(ws, row_index, columns, "unit_price")
    quantity = _to_int(_cell(ws, row_index, columns.get("quantity"))) or 1
    amount = _first_float_in_columns(ws, row_index, columns, "amount")
    if unit_price is None and amount is not None and quantity:
        unit_price = amount / quantity
    has_price = unit_price is not None and unit_price > 0
    final_price_cny = _convert_to_cny(unit_price, currency, usd_to_cny) if has_price else None
    width = _to_float(_cell(ws, row_index, columns.get("width_cm")))
    depth = _to_float(_cell(ws, row_index, columns.get("depth_cm")))
    height = _to_float(_cell(ws, row_index, columns.get("height_cm")))
    dimension_text = _clean_text(_cell(ws, row_index, columns.get("dimension_text")))
    if (
        material
        and dimension_text
        and _looks_like_dimension_value(material)
        and not _looks_like_dimension_value(dimension_text)
        and _looks_like_material_value(dimension_text)
    ):
        material, dimension_text = dimension_text, material
    elif material and not dimension_text and _looks_like_dimension_value(material):
        dimension_text, material = material, ""
    parsed_dimensions = _dimension_payload_from_text(dimension_text)
    dimensions_cm = _dimensions_cm(width, depth, height)
    dimensions_mm = _dimensions_mm_from_cm(width, depth, height)
    if parsed_dimensions and not _has_full_dimensions(dimensions_cm):
        dimensions_cm = parsed_dimensions["dimensions_cm"]
        dimensions_mm = parsed_dimensions["dimensions_mm"]
    cbm = _to_float(_cell(ws, row_index, columns.get("cbm")))
    dimension_display_text = (
        parsed_dimensions["raw_text"]
        if parsed_dimensions else
        (dimension_text if columns.get("dimension_text") else "")
    )
    if not _looks_like_product_row(product_name, item_no, material, dimension_display_text, unit_price, amount):
        return None

    description_parts = [product_name, item_no, material, finish, fabric, requirements]
    return {
        "product_name": product_name or category_text or item_no or "未命名产品",
        "item_no": item_no,
        "category": category,
        "material": material,
        "finish": finish,
        "fabric": fabric,
        "top": _clean_text(_cell(ws, row_index, columns.get("top"))),
        "glass": _clean_text(_cell(ws, row_index, columns.get("glass"))),
        "rail": _clean_text(_cell(ws, row_index, columns.get("rail"))),
        "veneer": _clean_text(_cell(ws, row_index, columns.get("veneer"))),
        "requirements": requirements,
        "dimensions_cm": dimensions_cm,
        "dimensions_mm": dimensions_mm,
        "dimension_text": dimension_display_text,
        "dimension_unit": parsed_dimensions["source_unit"] if parsed_dimensions else ("cm" if dimensions_cm else ""),
        "cbm": cbm,
        "quantity": quantity,
        "unit": unit_text,
        "original_unit_price": unit_price,
        "original_currency": currency,
        "unit_price_cny": final_price_cny,
        "has_price": has_price,
        "amount_original": amount,
        "description": " ".join(part for part in description_parts if part),
        "raw_row": {key: _clean_text(value) for key, value in raw_values.items()},
    }


def _row_image_paths(images: list[QuoteImportImage], columns: dict[str, Any]) -> list[str]:
    if not images:
        return []
    image_col = columns.get("image_col")
    if image_col:
        filtered = [image for image in images if abs(image.anchor_col - image_col) <= 1]
        if filtered:
            return [image.file_path for image in filtered]
    return [image.file_path for image in images]


def _looks_like_product_row(
    product_name: str,
    item_no: str,
    material: str,
    dimension_text: str,
    unit_price: float | None,
    amount: float | None,
) -> bool:
    identity = product_name or item_no
    evidence = material or dimension_text or unit_price or amount
    if not identity or not evidence:
        return False
    if re.search(r"合计|小计|总计|备注|单位|数量|单价|样板间|家具选型", identity):
        return False
    return True


def _item_has_price(item: dict[str, Any]) -> bool:
    price = _to_float(item.get("unit_price_cny"))
    return price is not None and price > 0


def _first_float_in_columns(ws, row_index: int, columns: dict[str, Any], key: str) -> float | None:
    candidates = list(columns.get(f"{key}_candidates") or [])
    first_col = columns.get(key)
    if first_col and first_col not in candidates:
        candidates.insert(0, first_col)
    for col in candidates:
        number = _to_float(_cell(ws, row_index, col))
        if number is not None:
            return number
    return None


def _line_item(line: str, index: int, meta: dict[str, Any], usd_to_cny: float) -> dict[str, Any] | None:
    text = line.strip()
    if not text or _is_noise_text(text):
        return None
    price_match = re.search(r"(USD|\$|RMB|CNY|￥)?\s*([0-9][0-9,]*(?:\.[0-9]+)?)", text, re.I)
    parsed_dimensions = _dimension_payload_from_text(text, require_marker=True)
    if not price_match or not parsed_dimensions:
        return None
    price = _to_float(price_match.group(2))
    if price is None or price <= 0:
        return None
    currency = "USD" if re.search(r"USD|\$", price_match.group(0), re.I) else meta.get("currency", "CNY")
    cleaned_name = text[:price_match.start()].strip(" -:：")
    return {
        "product_name": cleaned_name or f"PDF产品{index}",
        "item_no": "",
        "category": _category_code(cleaned_name, ""),
        "material": "",
        "finish": "",
        "fabric": "",
        "requirements": text,
        "dimensions_cm": parsed_dimensions["dimensions_cm"],
        "dimensions_mm": parsed_dimensions["dimensions_mm"],
        "dimension_text": parsed_dimensions["raw_text"],
        "dimension_unit": parsed_dimensions["source_unit"],
        "quantity": 1,
        "original_unit_price": price,
        "original_currency": currency,
        "unit_price_cny": _convert_to_cny(price, currency, usd_to_cny),
        "amount_original": None,
        "description": text,
        "raw_row": {"text": text},
    }


def _item_to_quote(
    item: dict[str, Any],
    meta: dict[str, Any],
    usd_to_cny: float,
    request: QuoteImportRequest,
) -> HistoricalQuoteCreate:
    if not _item_has_price(item):
        raise ValueError("缺少单价或合计，作为待补价预览行处理")
    quote_no = str(meta.get("quote_no") or "QUOTE")
    row_index = int(item.get("row_index") or 0)
    category = item.get("category") or "CH"
    product_id = _product_id(item, category, meta, row_index)
    quote_id = _quote_id(quote_no, product_id, row_index)
    features = {
        "category": category,
        "item_no": item.get("item_no", ""),
        "source_label": request.source_label,
        "source_file": meta.get("source_file", ""),
        "source_sheet": item.get("sheet_name", ""),
        "source_row": row_index,
        "source_quote_no": quote_no,
        "customer_name": meta.get("customer_name", ""),
        "dimensions_cm": item.get("dimensions_cm", {}),
        "dimensions_mm": item.get("dimensions_mm", {}),
        "dimension_text": item.get("dimension_text", ""),
        "dimension_unit": item.get("dimension_unit", ""),
        "cbm": item.get("cbm"),
        "material": item.get("material", ""),
        "finish": item.get("finish", ""),
        "fabric": item.get("fabric", ""),
        "top": item.get("top", ""),
        "glass": item.get("glass", ""),
        "rail": item.get("rail", ""),
        "veneer": item.get("veneer", ""),
        "requirements": item.get("requirements", ""),
        "unit": item.get("unit", ""),
        "image_paths": item.get("image_paths", []),
        "original_currency": item.get("original_currency", ""),
        "original_unit_price": item.get("original_unit_price"),
        "exchange_rate_to_cny": usd_to_cny if item.get("original_currency") == "USD" else 1.0,
        "unit_price_cny": item.get("unit_price_cny"),
        "amount_original": item.get("amount_original"),
        "price_status": "priced",
        "raw_row": item.get("raw_row", {}),
        "llm": item.get("llm", {}),
    }
    return HistoricalQuoteCreate(
        quote_id=quote_id,
        product_id=product_id,
        product_name=str(item.get("product_name") or product_id),
        category=category,
        description=str(item.get("description") or item.get("requirements") or ""),
        features={key: value for key, value in features.items() if value not in (None, "", [], {})},
        customer_tier=request.default_customer_tier,
        region=request.default_region or ("US" if item.get("original_currency") == "USD" else "CN"),
        quantity=int(item.get("quantity") or 1),
        final_price_cny=round(float(item["unit_price_cny"]), 2),
        quoted_at=str(meta.get("quote_date") or _date_from_filename(str(meta.get("source_file", ""))) or date.today().isoformat()),
    )


def _item_to_preview_quote(
    item: dict[str, Any],
    meta: dict[str, Any],
    request: QuoteImportRequest,
) -> dict[str, Any]:
    quote_no = str(meta.get("quote_no") or "QUOTE")
    row_index = int(item.get("row_index") or 0)
    category = item.get("category") or "CH"
    product_id = _product_id(item, category, meta, row_index)
    quote_id = f"PREVIEW-{_quote_id(quote_no, product_id, row_index)}"
    features = {
        "category": category,
        "item_no": item.get("item_no", ""),
        "source_label": request.source_label,
        "source_file": meta.get("source_file", ""),
        "source_sheet": item.get("sheet_name", ""),
        "source_row": row_index,
        "source_quote_no": quote_no,
        "customer_name": meta.get("customer_name", ""),
        "dimensions_cm": item.get("dimensions_cm", {}),
        "dimensions_mm": item.get("dimensions_mm", {}),
        "dimension_text": item.get("dimension_text", ""),
        "dimension_unit": item.get("dimension_unit", ""),
        "cbm": item.get("cbm"),
        "material": item.get("material", ""),
        "finish": item.get("finish", ""),
        "fabric": item.get("fabric", ""),
        "top": item.get("top", ""),
        "glass": item.get("glass", ""),
        "rail": item.get("rail", ""),
        "veneer": item.get("veneer", ""),
        "requirements": item.get("requirements", ""),
        "unit": item.get("unit", ""),
        "image_paths": item.get("image_paths", []),
        "original_currency": item.get("original_currency", ""),
        "original_unit_price": item.get("original_unit_price"),
        "unit_price_cny": item.get("unit_price_cny"),
        "amount_original": item.get("amount_original"),
        "price_status": "missing_price",
        "raw_row": item.get("raw_row", {}),
        "llm": item.get("llm", {}),
    }
    now = datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    return {
        "quote_id": quote_id,
        "product_id": product_id,
        "product_name": str(item.get("product_name") or product_id),
        "category": category,
        "description": str(item.get("description") or item.get("requirements") or ""),
        "features": {key: value for key, value in features.items() if value not in (None, "", [], {})},
        "customer_tier": request.default_customer_tier,
        "region": request.default_region or "CN",
        "quantity": int(item.get("quantity") or 1),
        "final_price_cny": None,
        "quoted_at": str(meta.get("quote_date") or _date_from_filename(str(meta.get("source_file", ""))) or date.today().isoformat()),
        "created_at": now,
        "updated_at": now,
        "_previewOnly": True,
    }


def _preview_quote(data: HistoricalQuoteCreate) -> HistoricalQuote:
    now = datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    return HistoricalQuote(**data.model_dump(), created_at=now, updated_at=now)


def _extract_xlsx_images(path: Path, import_id: str, image_dir: Path, warnings: list[str]) -> list[QuoteImportImage]:
    images: list[QuoteImportImage] = []
    try:
        with zipfile.ZipFile(path) as archive:
            sheet_names = _xlsx_sheet_names(archive)
            for sheet_index, sheet_name in enumerate(sheet_names, start=1):
                rels_path = f"xl/worksheets/_rels/sheet{sheet_index}.xml.rels"
                if rels_path not in archive.namelist():
                    continue
                drawing_target = _worksheet_drawing_target(archive.read(rels_path))
                if not drawing_target:
                    continue
                drawing_path = _normalize_xlsx_path("xl/worksheets", drawing_target)
                drawing_rels = _drawing_relationships(archive, drawing_path)
                for seq, anchor in enumerate(_drawing_anchors(archive.read(drawing_path)), start=1):
                    target = drawing_rels.get(anchor["rid"])
                    if not target:
                        continue
                    media_path = _normalize_xlsx_path(str(Path(drawing_path).parent), target)
                    if media_path not in archive.namelist():
                        continue
                    suffix = Path(media_path).suffix or ".png"
                    safe_sheet = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff_-]+", "_", sheet_name)
                    out_dir = image_dir / import_id
                    out_dir.mkdir(parents=True, exist_ok=True)
                    out_path = out_dir / f"{safe_sheet}_R{anchor['row']}_C{anchor['col']}_{seq}{suffix}"
                    out_path.write_bytes(archive.read(media_path))
                    images.append(QuoteImportImage(
                        image_id=f"{import_id}-{sheet_index}-{seq}",
                        source_sheet=sheet_name,
                        anchor_row=int(anchor["row"]),
                        anchor_col=int(anchor["col"]),
                        file_path=str(out_path),
                    ))
    except Exception as exc:  # noqa: BLE001 图片抽取失败不应阻断价格导入
        warnings.append(f"图片抽取失败: {exc}")
    return images


def _xlsx_sheet_names(archive: zipfile.ZipFile) -> list[str]:
    workbook_xml = ET.fromstring(archive.read("xl/workbook.xml"))
    ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    names = [
        sheet.attrib.get("name", f"Sheet{idx}")
        for idx, sheet in enumerate(workbook_xml.findall(".//main:sheets/main:sheet", ns), start=1)
    ]
    return names or ["Sheet1"]


def _worksheet_drawing_target(rels_xml: bytes) -> str | None:
    root = ET.fromstring(rels_xml)
    for rel in root.findall("rel:Relationship", DRAWING_NS):
        if "drawing" in rel.attrib.get("Type", ""):
            return rel.attrib.get("Target")
    return None


def _drawing_relationships(archive: zipfile.ZipFile, drawing_path: str) -> dict[str, str]:
    rels_path = f"{Path(drawing_path).parent}/_rels/{Path(drawing_path).name}.rels"
    if rels_path not in archive.namelist():
        return {}
    root = ET.fromstring(archive.read(rels_path))
    return {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in root.findall("rel:Relationship", DRAWING_NS)
        if "image" in rel.attrib.get("Type", "")
    }


def _drawing_anchors(xml: bytes) -> list[dict[str, Any]]:
    root = ET.fromstring(xml)
    anchors: list[dict[str, Any]] = []
    for anchor in list(root.findall("xdr:twoCellAnchor", DRAWING_NS)) + list(root.findall("xdr:oneCellAnchor", DRAWING_NS)):
        from_node = anchor.find("xdr:from", DRAWING_NS)
        blip = anchor.find(".//a:blip", DRAWING_NS)
        if from_node is None or blip is None:
            continue
        rid = blip.attrib.get(f"{{{DRAWING_NS['r']}}}embed")
        row_node = from_node.find("xdr:row", DRAWING_NS)
        col_node = from_node.find("xdr:col", DRAWING_NS)
        if not rid or row_node is None or col_node is None:
            continue
        anchors.append({"rid": rid, "row": int(row_node.text or "0") + 1, "col": int(col_node.text or "0") + 1})
    return anchors


def _normalize_xlsx_path(base_dir: str, target: str) -> str:
    target_path = Path(base_dir) / target
    parts: list[str] = []
    for part in target_path.as_posix().split("/"):
        if part in ("", "."):
            continue
        if part == "..":
            if parts:
                parts.pop()
            continue
        parts.append(part)
    return "/".join(parts)


def _maybe_enrich_with_llm(
    items: list[dict[str, Any]],
    meta: dict[str, Any],
    request: QuoteImportRequest,
    cfg: dict[str, Any],
    warnings: list[str],
) -> None:
    llm_cfg = cfg.get("llm", {}) if isinstance(cfg.get("llm"), dict) else {}
    use_llm = request.use_llm if request.use_llm is not None else bool(llm_cfg.get("enabled_by_default", False))
    if not use_llm or not items:
        return
    api_key_env = str(llm_cfg.get("api_key_env") or "DASHSCOPE_API_KEY")
    if not os.environ.get(api_key_env):
        load_from_keychain([api_key_env])
    api_key = os.environ.get(api_key_env)
    if not api_key:
        warnings.append(f"未找到 {api_key_env}，已跳过大模型辅助解析")
        return
    base_url = str(llm_cfg.get("base_url") or "https://dashscope.aliyuncs.com/compatible-mode/v1").rstrip("/")
    model = request.llm_model or llm_cfg.get("model") or "qwen3-max"
    rows = [_sanitize_llm_row(item) for item in items[:80]]
    prompt = (
        "你是家具历史报价单结构化助手。请只基于输入 JSON 行项目，补全 category/material/finish/"
        "requirements/style/craft/price_factors 字段，不要编造价格。输出严格 JSON："
        "{\"items\":[{\"row_index\":数字,\"patch\":{...}}]}。"
    )
    try:
        response = httpx.post(
            f"{base_url}/chat/completions",
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json={
                "model": model,
                "temperature": float(llm_cfg.get("temperature", 0.1)),
                "max_tokens": int(llm_cfg.get("max_tokens", 3000)),
                "messages": [
                    {"role": "system", "content": prompt},
                    {"role": "user", "content": json.dumps({"meta": _safe_meta(meta), "rows": rows}, ensure_ascii=False)},
                ],
            },
            timeout=30,
        )
        response.raise_for_status()
        content = response.json()["choices"][0]["message"]["content"]
        payload = _json_from_text(content)
        patches = {int(row["row_index"]): row.get("patch", {}) for row in payload.get("items", []) if "row_index" in row}
        for item in items:
            patch = patches.get(int(item.get("row_index") or 0))
            if patch:
                item.update({key: value for key, value in patch.items() if value not in (None, "", [], {})})
                item["llm"] = {"provider": request.llm_provider_id or llm_cfg.get("provider_id", "dashscope"), "model": model}
    except Exception as exc:  # noqa: BLE001 外部模型失败时保留本地解析结果
        warnings.append(f"大模型辅助解析失败，已使用本地解析结果: {exc}")


def _analysis_factors(quote: HistoricalQuote) -> dict[str, Any]:
    features = quote.features or {}
    dims = features.get("dimensions_cm") or {}
    width = _to_float(dims.get("width"))
    depth = _to_float(dims.get("depth"))
    height = _to_float(dims.get("height"))
    return {
        "category": quote.category,
        "material": features.get("material"),
        "finish": features.get("finish"),
        "fabric": features.get("fabric"),
        "top": features.get("top"),
        "customer_tier": quote.customer_tier,
        "region": quote.region,
        "quantity_bucket": _quantity_bucket(quote.quantity),
        "width_bucket_cm": _bucket(width, 50) if width else "",
        "height_bucket_cm": _bucket(height, 50) if height else "",
        "volume_bucket": _volume_bucket(width, depth, height),
    }


def _sanitize_llm_row(item: dict[str, Any]) -> dict[str, Any]:
    raw = item.get("raw_row", {})
    safe_raw = {key: value for key, value in raw.items() if not _is_noise_text(str(value))}
    return {
        "row_index": item.get("row_index"),
        "product_name": item.get("product_name"),
        "item_no": item.get("item_no"),
        "category": item.get("category"),
        "description": item.get("description"),
        "material": item.get("material"),
        "finish": item.get("finish"),
        "fabric": item.get("fabric"),
        "requirements": item.get("requirements"),
        "dimensions_cm": item.get("dimensions_cm"),
        "original_currency": item.get("original_currency"),
        "original_unit_price": item.get("original_unit_price"),
        "raw_row": safe_raw,
    }


def _safe_meta(meta: dict[str, Any]) -> dict[str, Any]:
    return {key: value for key, value in meta.items() if key not in {"raw_text_preview"} and not _is_noise_text(str(value))}


def _json_from_text(text: str) -> dict[str, Any]:
    clean = text.strip()
    if clean.startswith("```"):
        clean = re.sub(r"^```(?:json)?|```$", "", clean, flags=re.I | re.M).strip()
    start = clean.find("{")
    end = clean.rfind("}")
    if start >= 0 and end >= start:
        clean = clean[start:end + 1]
    return json.loads(clean)


def _header_text(ws, row: int, col: int) -> str:
    current = _clean_text(ws.cell(row, col).value)
    previous = _clean_text(ws.cell(row - 1, col).value) if row > 1 else ""
    if current and re.fullmatch(r"USD|CNY|RMB|\$", current, re.I):
        return " ".join(part for part in [previous, current] if part)
    return current or previous


def _classify_header(text: str) -> str | None:
    norm = _norm(text)
    raw = text.lower()
    if re.search(r"amount|总价|合计", raw):
        return "amount"
    if re.search(r"exworks|ex works|单价|unitprice|unit price", raw) or (re.search(r"\busd\b", raw) and re.search(r"exworks|price|单价", raw)):
        return "unit_price"
    if re.search(r"\bqty\b|数量", raw):
        return "quantity"
    if re.search(r"单位|unit\b", raw):
        return "unit"
    if re.search(r"图片|picture|image|photo", raw):
        return "image_col"
    if re.search(r"item\s*no|sku|型号|款号", raw) or "itemno" in norm:
        return "item_no"
    if re.search(r"产品名称|productname|名称|内容", raw):
        return "product_name"
    if re.search(r"规格|尺寸|specification|dimensions?", raw):
        return "dimension_text"
    if re.search(r"category|品类|类别", raw):
        return "category_text"
    if re.search(r"materials?|材质", raw):
        return "material"
    if re.search(r"finish\s*/\s*description|涂装|油漆|finish|颜色|color", raw):
        return "finish"
    if re.search(r"fabric|面料", raw):
        return "fabric"
    if re.search(r"remarks?|question|备注|要求", raw):
        return "requirements"
    if re.search(r"width|\b宽\b|宽度", raw):
        return "width_cm"
    if re.search(r"depth|\b深\b|深度", raw):
        return "depth_cm"
    if re.search(r"height|\b高\b|高度", raw):
        return "height_cm"
    if re.search(r"\bcbm\b|体积", raw):
        return "cbm"
    if re.search(r"\btop\b|台面|桌面", raw):
        return "top"
    if re.search(r"glass|玻璃", raw):
        return "glass"
    if re.search(r"rail|滑轨", raw):
        return "rail"
    if re.search(r"veneer|木皮", raw):
        return "veneer"
    if re.search(r"交货|lead", raw):
        return "lead_time"
    return None


def _is_dimension_group(text: str) -> bool:
    return bool(re.search(r"W\s*[-/]\s*D\s*[-/]\s*H|常规尺寸|尺寸", text, re.I))


def _is_split_dimension_group(text: str) -> bool:
    return bool(re.search(r"W\s*[-/]\s*D\s*[-/]\s*H|常规尺寸", text, re.I))


def _looks_like_dimension_value(text: str) -> bool:
    clean = _clean_text(text)
    if not clean:
        return False
    return bool(re.search(
        r"常规尺寸|尺寸待定|[WDHL]\s*:?\s*\d|[宽深高长]\s*:?\s*\d|"
        r"\d+(?:\.\d+)?\s*(?:mm|毫米|cm|厘米|公分|米)\b|[*x×＊/]",
        clean,
        re.I,
    ))


def _looks_like_material_value(text: str) -> bool:
    clean = _clean_text(text)
    if not clean:
        return False
    return bool(re.search(
        r"木|实木|板|漆|金属|五金|布艺|面料|皮|海绵|玻璃|石|大理石|玉石|雕|拉手|水性|环保|框架",
        clean,
        re.I,
    ))


def _cell(ws, row: int, col: int | None) -> Any:
    if not col:
        return None
    if col < 1 or col > ws.max_column:
        return None
    return ws.cell(row, col).value


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    return re.sub(r"\s+", " ", str(value).replace("\n", " ")).strip()


def _norm(text: str) -> str:
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", text.lower())


def _to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value)
    if text.startswith("="):
        return None
    match = re.search(r"-?\d+(?:,\d{3})*(?:\.\d+)?|-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    return float(match.group(0).replace(",", ""))


def _to_int(value: Any) -> int | None:
    number = _to_float(value)
    return int(number) if number is not None else None


def _dimensions_cm(width: float | None, depth: float | None, height: float | None) -> dict[str, float]:
    result = {}
    if width is not None:
        result["width"] = round(width, 2)
    if depth is not None:
        result["depth"] = round(depth, 2)
    if height is not None:
        result["height"] = round(height, 2)
    return result


def _dimensions_mm_from_cm(width: float | None, depth: float | None, height: float | None) -> dict[str, float]:
    result = {}
    if width is not None:
        result["width"] = round(width * 10, 2)
    if depth is not None:
        result["depth"] = round(depth * 10, 2)
    if height is not None:
        result["height"] = round(height * 10, 2)
    return result


def _dimensions_mm(width: float | None, depth: float | None, height: float | None) -> dict[str, float]:
    return _dimensions_mm_from_cm(width, depth, height)


def _has_full_dimensions(dimensions: dict[str, float]) -> bool:
    return all(dimensions.get(key) for key in ("width", "depth", "height"))


def _dimension_payload_from_text(value: Any, require_marker: bool = False) -> dict[str, Any] | None:
    raw = _clean_text(value)
    if not raw:
        return None

    normalized = (
        raw.replace("×", "*")
        .replace("＊", "*")
        .replace("X", "*")
        .replace("：", ":")
    )
    has_marker = bool(re.search(r"[WDHL]\s*:?\s*\d|[宽深高长]\s*:?\s*\d|\d+\s*(?:mm|毫米|cm|厘米|公分)\b|[*x/]", normalized, re.I))
    if require_marker and not has_marker:
        return None

    labeled: dict[str, float] = {}
    for match in re.finditer(r"(?i)([WDHL宽深高长])\s*:?\s*([0-9][0-9,]*(?:\.[0-9]+)?)", normalized):
        label = match.group(1).upper()
        number = float(match.group(2).replace(",", ""))
        if label in {"W", "L", "宽", "长"}:
            labeled.setdefault("width", number)
        elif label in {"D", "深"}:
            labeled.setdefault("depth", number)
        elif label in {"H", "高"}:
            labeled.setdefault("height", number)

    if all(key in labeled for key in ("width", "depth", "height")):
        values = [labeled["width"], labeled["depth"], labeled["height"]]
    else:
        values = [
            float(match.group(0).replace(",", ""))
            for match in re.finditer(r"[0-9][0-9,]*(?:\.[0-9]+)?", normalized)
        ][:3]
        diameter_values = _diameter_dimension_values(normalized, labeled, values)
        if diameter_values:
            values = diameter_values

    if len(values) < 3:
        return None

    source_unit = _dimension_source_unit(raw, values)
    if source_unit == "mm":
        mm_values = values
        cm_values = [value / 10 for value in values]
    else:
        cm_values = values
        mm_values = [value * 10 for value in values]

    return {
        "raw_text": raw,
        "source_unit": source_unit,
        "dimensions_cm": _dimension_values(cm_values),
        "dimensions_mm": _dimension_values(mm_values),
    }


def _diameter_dimension_values(
    normalized: str,
    labeled: dict[str, float],
    numeric_values: list[float],
) -> list[float] | None:
    match = re.search(r"(?i)(?:[φΦØø]|直径|直徑|diameter|dia\.?)\s*([0-9][0-9,]*(?:\.[0-9]+)?)", normalized)
    if not match:
        return None
    diameter = float(match.group(1).replace(",", ""))
    height = labeled.get("height")
    if height is None and len(numeric_values) >= 2:
        height = numeric_values[1]
    if height is None:
        return None
    return [diameter, diameter, height]


def _dimension_source_unit(raw: str, values: list[float]) -> str:
    if re.search(r"mm|毫米", raw, re.I):
        return "mm"
    if re.search(r"cm|厘米|公分", raw, re.I):
        return "cm"
    return "mm" if max(values) > 300 else "cm"


def _dimension_values(values: list[float]) -> dict[str, float]:
    return {
        "width": round(values[0], 2),
        "depth": round(values[1], 2),
        "height": round(values[2], 2),
    }


def _convert_to_cny(price: float, currency: str, usd_to_cny: float) -> float:
    return round(price * usd_to_cny, 2) if currency == "USD" else round(price, 2)


def _detect_currency(text: str, fallback: str | None) -> str | None:
    if re.search(r"\bUSD\b|美元|\$", text, re.I):
        return "USD"
    if re.search(r"￥|RMB|CNY|人民币|不含税金", text, re.I):
        return "CNY"
    return fallback


def _extract_quote_no(text: str) -> str:
    patterns = [
        r"销售合同号[:：]?\s*([A-Z0-9_-]{5,})",
        r"\b(?:NO|No|PI|Quote)\b\s*[:：-]?\s*([A-Z]{0,4}[-A-Z0-9]{5,})",
        r"\b(PI[-_][A-Z0-9-]{3,}|PI[A-Z]{0,3}\d{4,})\b",
        r"\b(EA\d{2}[A-Z]{2}\d{4,})\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.I)
        if match:
            return match.group(1).strip()
    return ""


def _looks_like_quote_no(text: str) -> bool:
    clean = text.strip()
    if not clean:
        return False
    return bool(re.search(r"\d", clean) and re.search(r"[A-Z]", clean, re.I) and len(clean) >= 5)


def _extract_date(value: Any, text: str) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    patterns = [
        r"(20\d{2})[./-](\d{1,2})[./-](\d{1,2})",
        r"(20\d{2})年(\d{1,2})月(\d{1,2})日?",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            year, month, day = (int(part) for part in match.groups())
            return date(year, month, day).isoformat()
    return ""


def _date_from_filename(name: str) -> str:
    return _extract_date("", name)


def _extract_customer(text: str) -> str:
    patterns = [
        r"客户名称[:：]\s*([^;，,]+)",
        r"\bTo\s*[:：]\s*([A-Za-z0-9 ._-]+)",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.I)
        if match:
            return match.group(1).strip()
    return ""


def _category_code(text: str, item_no: str) -> str:
    value = f"{text} {item_no}".lower()
    rules = [
        ("SF", r"sofa|沙发"),
        ("BD", r"bed|床架|床\b"),
        ("CB", r"cabinet|柜|床头柜|nightstand|dresser|bookcase"),
        ("TB", r"table|desk|桌|台|边几|茶几"),
        ("CH", r"chair|椅|stool|凳"),
    ]
    for code, pattern in rules:
        if re.search(pattern, value):
            return code
    prefix = item_no[:2].upper()
    if prefix in {"SF", "BD", "CB", "TB", "CH"}:
        return prefix
    return "CH"


def _product_id(item: dict[str, Any], category: str, meta: dict[str, Any], row_index: int) -> str:
    item_no = str(item.get("item_no") or "").strip()
    if item_no:
        return item_no
    year = str(meta.get("quote_date") or "")[:4] or str(date.today().year)
    return f"{category}-{year}-{row_index:03d}"


def _quote_id(quote_no: str, product_id: str, row_index: int) -> str:
    base = re.sub(r"[^A-Za-z0-9_-]+", "-", quote_no).strip("-") or "QUOTE"
    product = re.sub(r"[^A-Za-z0-9_-]+", "-", product_id).strip("-") or f"R{row_index}"
    return f"{base}-{product}-R{row_index}"


def _import_id(path: Path) -> str:
    stem = re.sub(r"[^A-Za-z0-9\u4e00-\u9fff_-]+", "-", path.stem).strip("-")
    return f"IMP-{stem[:40]}-{uuid.uuid4().hex[:6].upper()}"


def _resolve_path(value: str | Path) -> Path:
    path = Path(value).expanduser()
    if not path.is_absolute():
        path = ROOT / path
    return path


def _quote_files(folder: Path, recursive: bool) -> list[Path]:
    pattern = "**/*" if recursive else "*"
    files = [
        path
        for path in folder.glob(pattern)
        if path.is_file()
        and path.suffix.lower() in {".xlsx", ".xlsm", ".pdf"}
        and not path.name.startswith(("~$", ".~"))
    ]
    return sorted(files, key=lambda item: str(item).lower())


def _safe_relative_path(value: str) -> Path:
    parts = []
    for part in Path(value).parts:
        clean = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff._ -]+", "_", part).strip(" .")
        if clean and clean not in {"..", "."}:
            parts.append(clean)
    if not parts:
        parts = [f"quote-{uuid.uuid4().hex[:8]}.xlsx"]
    return Path(*parts)


def _avg(values: list[float]) -> float:
    return sum(values) / len(values) if values else 0


def _quantity_bucket(quantity: int) -> str:
    if quantity >= 100:
        return "100+"
    if quantity >= 50:
        return "50-99"
    if quantity >= 10:
        return "10-49"
    if quantity >= 3:
        return "3-9"
    return "1-2"


def _bucket(value: float | None, step: int) -> str:
    if value is None:
        return ""
    low = int(value // step * step)
    return f"{low}-{low + step}"


def _volume_bucket(width: float | None, depth: float | None, height: float | None) -> str:
    if not width or not depth or not height:
        return ""
    volume = width * depth * height / 1_000_000
    if volume >= 2:
        return "2.0cbm+"
    if volume >= 1:
        return "1.0-2.0cbm"
    if volume >= 0.5:
        return "0.5-1.0cbm"
    return "<0.5cbm"


def _is_noise_text(text: str) -> bool:
    return bool(re.search(r"bank|swift|beneficiary|account|email|tel|phone|address|payment|wire|银行|账号|电话|邮箱|地址|付款", text, re.I))
