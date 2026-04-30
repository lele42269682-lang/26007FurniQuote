"""历史报价单导入测试。"""
from __future__ import annotations

import sys
from base64 import b64encode
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(ROOT))


@pytest.fixture()
def temp_db(tmp_path, monkeypatch):
    db_path = tmp_path / "furniquote_test.db"
    monkeypatch.setenv("DATABASE_URL", f"sqlite:///{db_path}")
    return db_path


def _save_cn_quote(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "客户报价单"
    ws["A1"] = "EUROANTIQUE 越然家居报价清单 NO：YRBJ001"
    ws["A2"] = "客户名称:厦门阮阮"
    ws["K2"] = "日期：2026.4.17"
    ws["N2"] = "销售合同号：EA26XS0087"
    headers = ["序号", "产品名称", "ITEM NO.", "常规尺寸（CM） / W - D - H", "", "", "CBM", "涂装 颜色", "COLOR", "面料代号", "材质", "图片", "备注", "单价", "数量", "总价", "交货周期"]
    ws.append([])
    ws.append(headers)
    ws.append([1, "床头柜", "NST380", 54, 45, 51, None, "墨蓝色", "", "", "桃花芯木框架；美国宣伟油漆；东泰托底滑轨；", "", "按照图纸生产", 1450, 2, "=N4*O4", "付款后90天"])
    wb.save(path)


def _save_usd_quote(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "1"
    ws["B8"] = "No :"
    ws["C8"] = "PI-EA260452"
    ws["B9"] = "Date :"
    ws["C9"] = datetime(2026, 4, 24)
    ws["B10"] = "To :"
    ws["C10"] = "Mr. Mark"
    ws["B11"] = "图片"
    ws["C11"] = "名称"
    ws["V11"] = "ExWorks"
    ws["W11"] = "Amount"
    headers = ["NO", "Picture", "Category", "materials", "SKU", "Width(cm)", "Depth", "Height", "CBM", "Total CBM", "TOP", "Picture", "Fabric", "Glass", "Rail", "Veneer Picture", "Finish Picture", "Finish / Description", "Finish", "Packages", "QTY", "USD", "USD", "Remarks/Question"]
    for idx, value in enumerate(headers, start=1):
        ws.cell(12, idx, value)
    row = [1, "", "CABINET", "MAHOGHANY WOOD", "", 190, 48, 95, 1.11, "=I13*U13", "MARBLE TOP", "", "", "", "", "", "", "SILVER LEAF", "", "", 1, 1230, "=U13*V13", "custom marble cabinet"]
    for idx, value in enumerate(row, start=1):
        ws.cell(13, idx, value)
    wb.save(path)


def _save_mm_spec_quote(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "询价单"
    headers = ["序号", "摆放区域", "产品图片", "产品名称", "规格", "材质", "单价（RMB）", "数量", "单位", "总价（RMB）", "备注"]
    ws.append(headers)
    ws.append([1, "", "", "多人沙发", "W2600*D1049*H1059", "木质+布艺", 10875, 1, "个", 10875, "7个抱枕"])
    ws.append([2, "", "", "长凳", "1200*D420*H480", "木质+布艺", 2725, 1, "个", 2725, ""])
    ws.append([3, "", "", "玄关桌", "换成了这个尺寸\nW1400*D381*H850", "木质", "", 1, "个", 5350, ""])
    ws.append([4, "", "", "书桌", "W1460*600*720", "木质", "", 1, "个", 4890, ""])
    wb.save(path)


def _save_selection_without_price_quote(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "家具"
    ws.append(["186顶复样板间家具选型概念方案示意图"])
    ws.append([])
    ws.append(["序号", "平面位置示意", "位置", "", "内容", "选型示意图片", "尺寸", "材质/颜色", "单位", "数量", "单价", "合计", "备注"])
    ws.append(["样板间--家具"])
    ws.append([1, "", "一层", "客厅", "三人位沙发", "", "3600*950*850", "桦木实木框架+实木多层板+高密度海绵+优质布艺+环保水性漆", "件", 1, "", "", ""])
    ws.append([2, "", "", "", "茶几", "", "1.35", "桦木实木框架+实木多层板+环保水性漆", "件", 1, "", "", ""])
    ws.append([3, "", "", "", "餐边柜", "", "1600*550", "桦木实木框架+实木多层板+机雕+环保水性漆+五金拉手", "件", 1, "", "", ""])
    ws.append([None, "", "", "书房", "单椅", "", "实木+布艺", "常规尺寸", "件", 2, "", "", ""])
    ws.append([None, "", "", "书房", "边柜", "", "", "w1300", "件", 1, "", "", ""])
    wb.save(path)


def _save_second_price_column_quote(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "询价表"
    ws.append(["报价"])
    ws.append(["CODE", "DESCRIPTION", "", "FURNITURE", "", "", "DESCRIPTION", "", "", "", "QTY.", "", "DESCRIPTION", "", "", "SUPPLIER", "", "", "", ""])
    ws.append(["编号", "名称", "", "图样", "", "", "规格说明(mm)", "", "", "", "数量", "", "材质说明", "", "", "材质说明", "单价\n含面料", "单价\n不含面料", "总价\n不含面料", "周期"])
    ws.append(["FF-100", "多人大沙发", "", "", "", "", "W2800mmxD950mmxH1200mm", "", "", "", 1, "", "榉木实木框架+布艺", "", "", "桃花芯实木框架", "", 10500, "=K4*R4", ""])
    ws.append(["FF-102", "茶几", "", "", "", "", "φ1200mmxH480mm", "", "", "", 1, "", "榉木实木框架+白色玉石台面", "", "", "桃花芯实木框架", "", 5980, "=K5*R5", ""])
    wb.save(path)


def test_import_cn_quote_file_preserves_dimensions_materials_and_price(temp_db, tmp_path):
    from modules.m08_stats.main import run

    file_path = tmp_path / "EA26XS0087厦门阮阮报价单20260416.xlsx"
    _save_cn_quote(file_path)

    result = run({
        "action": "import_quote_file",
        "import_request": {
            "file_path": str(file_path),
            "default_customer_tier": "L1",
            "default_region": "CN",
        },
    })

    assert result["status"] == "ok"
    data = result["data"]
    assert data["quote_no"] == "EA26XS0087"
    assert data["customer_name"] == "厦门阮阮"
    assert data["imported_count"] == 1
    quote = data["quotes"][0]
    assert quote["product_id"] == "NST380"
    assert quote["final_price_cny"] == 1450
    assert quote["quantity"] == 2
    assert quote["features"]["original_currency"] == "CNY"
    assert quote["features"]["dimensions_mm"] == {"width": 540.0, "depth": 450.0, "height": 510.0}
    assert "桃花芯木" in quote["features"]["material"]
    assert "按照图纸生产" in quote["features"]["requirements"]


def test_import_usd_quote_file_converts_unit_price_to_cny_and_analyzes_factors(temp_db, tmp_path):
    from modules.m08_stats.main import run

    file_path = tmp_path / "2026-04-24 PI-EA260452_Mr. Mark.xlsx"
    _save_usd_quote(file_path)

    result = run({
        "action": "import_quote_file",
        "import_request": {
            "file_path": str(file_path),
            "default_customer_tier": "L2",
            "default_region": "US",
        },
    })

    assert result["status"] == "ok"
    quote = result["data"]["quotes"][0]
    assert result["data"]["quote_no"] == "PI-EA260452"
    assert result["data"]["customer_name"] == "Mr. Mark"
    assert quote["category"] == "CB"
    assert quote["final_price_cny"] == 8610
    assert quote["features"]["original_currency"] == "USD"
    assert quote["features"]["original_unit_price"] == 1230
    assert quote["features"]["exchange_rate_to_cny"] == 7
    assert quote["features"]["dimensions_cm"] == {"width": 190.0, "depth": 48.0, "height": 95.0}
    assert quote["features"]["top"] == "MARBLE TOP"

    analysis = run({"action": "analyze_price_factors", "min_support": 1, "limit": 10})
    assert analysis["status"] == "ok"
    factors = {(item["factor_key"], item["factor_value"]) for item in analysis["data"]["factors"]}
    assert ("category", "CB") in factors
    assert ("material", "MAHOGHANY WOOD") in factors


def test_import_mm_spec_column_parses_wdh_as_millimeters(temp_db, tmp_path):
    from modules.m08_stats.main import run

    file_path = tmp_path / "澳诺亚询价12.17.xlsx"
    _save_mm_spec_quote(file_path)

    result = run({
        "action": "import_quote_file",
        "import_request": {
            "file_path": str(file_path),
            "default_customer_tier": "L1",
            "default_region": "CN",
        },
    })

    assert result["status"] == "ok"
    quotes = result["data"]["quotes"]
    assert result["data"]["imported_count"] == 4
    sofa = quotes[0]
    assert sofa["product_name"] == "多人沙发"
    assert sofa["features"]["dimensions_mm"] == {"width": 2600.0, "depth": 1049.0, "height": 1059.0}
    assert sofa["features"]["dimensions_cm"] == {"width": 260.0, "depth": 104.9, "height": 105.9}
    assert sofa["features"]["dimension_unit"] == "mm"
    assert sofa["features"]["dimension_text"] == "W2600*D1049*H1059"
    bench = quotes[1]
    assert bench["features"]["dimensions_mm"] == {"width": 1200.0, "depth": 420.0, "height": 480.0}
    console = quotes[2]
    assert console["features"]["dimensions_mm"] == {"width": 1400.0, "depth": 381.0, "height": 850.0}
    desk = quotes[3]
    assert desk["features"]["dimensions_mm"] == {"width": 1460.0, "depth": 600.0, "height": 720.0}


def test_import_selection_file_without_prices_as_preview_items(temp_db, tmp_path):
    from modules.m08_stats.main import run

    file_path = tmp_path / "186顶复样板间家具含税清单空白20250702.xlsx"
    _save_selection_without_price_quote(file_path)

    result = run({
        "action": "import_quote_file",
        "import_request": {
            "file_path": str(file_path),
            "default_customer_tier": "L1",
            "default_region": "CN",
        },
    })

    assert result["status"] == "ok"
    data = result["data"]
    assert data["imported_count"] == 0
    assert data["preview_count"] == 5
    assert data["recognized_count"] == 5
    sofa = data["preview_items"][0]
    assert sofa["_previewOnly"] is True
    assert sofa["product_name"] == "三人位沙发"
    assert sofa["final_price_cny"] is None
    assert sofa["quantity"] == 1
    assert sofa["features"]["price_status"] == "missing_price"
    assert sofa["features"]["dimensions_mm"] == {"width": 3600.0, "depth": 950.0, "height": 850.0}
    assert "桦木实木框架" in sofa["features"]["material"]
    tea_table = data["preview_items"][1]
    assert tea_table["features"]["dimension_text"] == "1.35"
    chair = data["preview_items"][3]
    assert chair["product_name"] == "单椅"
    assert chair["quantity"] == 2
    assert chair["features"]["dimension_text"] == "常规尺寸"
    assert chair["features"]["material"] == "实木+布艺"
    sideboard = data["preview_items"][4]
    assert sideboard["product_name"] == "边柜"
    assert sideboard["features"]["dimension_text"] == "w1300"
    assert "material" not in sideboard["features"]


def test_import_quote_uses_later_unit_price_candidate_when_first_is_blank(temp_db, tmp_path):
    from modules.m08_stats.main import run

    file_path = tmp_path / "8 20250711古典定制家具询价.xlsx"
    _save_second_price_column_quote(file_path)

    result = run({
        "action": "import_quote_file",
        "import_request": {
            "file_path": str(file_path),
            "default_customer_tier": "L1",
            "default_region": "CN",
        },
    })

    assert result["status"] == "ok"
    data = result["data"]
    assert data["imported_count"] == 2
    assert data["preview_count"] == 0
    quote = data["quotes"][0]
    assert quote["product_name"] == "多人大沙发"
    assert quote["final_price_cny"] == 10500
    assert quote["quantity"] == 1
    assert quote["features"]["dimensions_mm"] == {"width": 2800.0, "depth": 950.0, "height": 1200.0}
    table = data["quotes"][1]
    assert table["product_name"] == "茶几"
    assert table["final_price_cny"] == 5980
    assert table["features"]["dimension_text"] == "φ1200mmxH480mm"
    assert table["features"]["dimensions_mm"] == {"width": 1200.0, "depth": 1200.0, "height": 480.0}


def test_import_uploaded_quote_file_from_browser_folder(temp_db, tmp_path):
    from modules.m08_stats.main import run

    file_path = tmp_path / "folder" / "EA26XS0087厦门阮阮报价单20260416.xlsx"
    file_path.parent.mkdir()
    _save_cn_quote(file_path)

    result = run({
        "action": "import_uploaded_quote_file",
        "upload_request": {
            "file_name": file_path.name,
            "relative_path": f"报价单/{file_path.name}",
            "content_base64": b64encode(file_path.read_bytes()).decode("ascii"),
            "default_customer_tier": "L1",
            "default_region": "CN",
        },
    })

    assert result["status"] == "ok"
    assert result["data"]["imported_count"] == 1
    quote = result["data"]["quotes"][0]
    assert quote["product_name"] == "床头柜"
    assert quote["features"]["dimensions_cm"] == {"width": 54.0, "depth": 45.0, "height": 51.0}
    assert quote["features"]["material"].startswith("桃花芯木框架")


def test_import_quote_folder_from_local_path_recursively(temp_db, tmp_path):
    from modules.m08_stats.main import run

    folder = tmp_path / "历史报价单" / "测试1"
    nested = folder / "子文件夹"
    nested.mkdir(parents=True)
    _save_cn_quote(folder / "EA26XS0087厦门阮阮报价单20260416.xlsx")
    (folder / ".~EA26XS0087厦门阮阮报价单20260416.xlsx").write_text("wps temp lock", encoding="utf-8")
    _save_usd_quote(nested / "2026-04-24 PI-EA260452_Mr. Mark.xlsx")

    result = run({
        "action": "import_quote_folder",
        "folder_request": {
            "folder_path": str(folder),
            "recursive": True,
            "default_customer_tier": "L1",
            "usd_to_cny": 7,
        },
    })

    assert result["status"] == "ok"
    data = result["data"]
    assert data["file_type"] == "folder"
    assert data["found_count"] == 2
    assert data["imported_count"] == 2
    assert len(data["files"]) == 2
    names = {quote["product_name"] for quote in data["quotes"]}
    assert names == {"床头柜", "CABINET"}
    assert all(quote["_sourceFile"] for quote in data["quotes"])
